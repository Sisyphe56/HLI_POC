#!/usr/bin/env python3
"""통합 상품 데이터 비교 스크립트.

Usage:
    python compare_product_data.py --data-set payment_cycle
    python compare_product_data.py --data-set annuity_age
"""
import argparse
import csv
import io
import json
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Dict, List, Optional, Set, Tuple
import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parent
DEFAULT_MAPPING_CSV = PROJECT_ROOT / 'config' / '보종코드_상품코드_매핑.csv'
DATASET_CONFIGS_PATH = PROJECT_ROOT / 'config' / 'dataset_configs.json'


# ──────────────────────────────────────────────
# 공통 유틸
# ──────────────────────────────────────────────

def normalize_text(value) -> str:
    if value is None or value == '' or (isinstance(value, float) and value != value):
        return ''
    s = unicodedata.normalize('NFKC', str(value).strip())
    s = s.replace('Ⅰ', 'I').replace('Ⅱ', 'II').replace('Ⅲ', 'III')
    s = s.replace('\u00a0', ' ')
    s = re.sub(r'^[\s·•▪∙◦]+\s*', '', s)
    s = re.sub(r'\s+', ' ', s)
    return s


def normalize_code(value) -> str:
    s = normalize_text(value)
    s = s.lstrip('0') or '0'
    return s


def read_text_with_fallback(path: Path) -> str:
    for enc in ('utf-8-sig', 'utf-8', 'cp949', 'euc-kr'):
        try:
            return path.read_text(encoding=enc)
        except UnicodeDecodeError:
            pass
    raise RuntimeError(f'Cannot decode {path}')


def load_csv_rows(csv_path: Path) -> List[dict]:
    """CSV의 모든 row를 로드 (261건, prod_dtcd/prod_itcd 포함)."""
    text = read_text_with_fallback(csv_path)
    rows = []
    for raw in csv.DictReader(io.StringIO(text)):
        dtcd = normalize_code(raw.get('ISRN_KIND_DTCD', ''))
        itcd = normalize_code(raw.get('ISRN_KIND_ITCD', ''))
        sale_nm = normalize_text(raw.get('ISRN_KIND_SALE_NM', ''))
        prod_dtcd = normalize_code(raw.get('PROD_DTCD', ''))
        prod_itcd = normalize_code(raw.get('PROD_ITCD', ''))
        prod_sale_nm = normalize_text(raw.get('PROD_SALE_NM', ''))
        if dtcd and itcd:
            rows.append({
                'isrn_kind_dtcd': dtcd,
                'isrn_kind_itcd': itcd,
                'isrn_kind_sale_nm': sale_nm,
                'prod_dtcd': prod_dtcd,
                'prod_itcd': prod_itcd,
                'prod_sale_nm': prod_sale_nm,
            })
    return rows


# ──────────────────────────────────────────────
# 공통 로더
# ──────────────────────────────────────────────

def load_answer_excel(
    excel_path: Path,
    key_cols: List[str],
    value_cols: List[str],
) -> Dict[Tuple[str, ...], List[dict]]:
    """정답 엑셀을 로드하여 key_cols 기준으로 그룹화.

    key_cols 중 ISRN_KIND_DTCD, ISRN_KIND_ITCD 는 normalize_code 적용,
    나머지는 normalize_text 적용.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers = [str(c.value) if c.value else '' for c in ws[1]]

    code_fields = {'ISRN_KIND_DTCD', 'ISRN_KIND_ITCD', 'PROD_DTCD', 'PROD_ITCD'}
    all_cols = set(key_cols) | set(value_cols)

    answer_data: Dict[Tuple[str, ...], List[dict]] = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw = {}
        for col_idx, value in enumerate(row):
            if col_idx < len(headers):
                raw[headers[col_idx]] = value

        # 정규화
        normed = {}
        for col in all_cols:
            v = raw.get(col, '')
            if col in code_fields:
                normed[col] = normalize_code(v)
            else:
                normed[col] = normalize_text(str(v))

        # key 생성
        key_vals = tuple(normed[c] for c in key_cols)
        if not all(key_vals):
            continue

        entry = {'row_num': row_idx}
        for c in all_cols:
            entry[c] = normed[c]

        answer_data.setdefault(key_vals, []).append(entry)

    wb.close()
    return answer_data


def load_answer_csv(
    csv_path: Path,
    key_cols: List[str],
    value_cols: List[str],
    encoding: str = 'euc-kr',
) -> Dict[Tuple[str, ...], List[dict]]:
    """정답 CSV를 로드하여 key_cols 기준으로 그룹화.

    load_answer_excel 과 동일한 정규화 로직 적용.
    """
    code_fields = {'ISRN_KIND_DTCD', 'ISRN_KIND_ITCD', 'PROD_DTCD', 'PROD_ITCD'}
    all_cols = set(key_cols) | set(value_cols)

    answer_data: Dict[Tuple[str, ...], List[dict]] = {}

    with csv_path.open('r', encoding=encoding) as f:
        reader = csv.DictReader(f)
        for row_idx, raw in enumerate(reader, start=2):
            normed = {}
            for col in all_cols:
                v = raw.get(col, '')
                if col in code_fields:
                    normed[col] = normalize_code(v)
                else:
                    normed[col] = normalize_text(str(v))

            key_vals = tuple(normed[c] for c in key_cols)
            if not all(key_vals):
                continue

            entry = {'row_num': row_idx}
            for c in all_cols:
                entry[c] = normed[c]

            answer_data.setdefault(key_vals, []).append(entry)

    return answer_data


def load_mapped_json_files(mapped_dir: Path, prefix: str) -> List[dict]:
    """매핑된 JSON 파일들을 로드."""
    all_rows = []
    for json_path in sorted(mapped_dir.glob(f'{prefix}*.json')):
        with json_path.open('r', encoding='utf-8') as f:
            data = json.load(f)
        if not isinstance(data, list):
            continue
        for row in data:
            if isinstance(row, dict):
                row['source_file'] = json_path.name
                all_rows.append(row)
    return all_rows


# ──────────────────────────────────────────────
# dataset config 로드
# ──────────────────────────────────────────────

def _load_dataset_configs() -> dict:
    with DATASET_CONFIGS_PATH.open('r', encoding='utf-8') as f:
        raw = json.load(f)
    # _설명 같은 메타 키 제거
    return {k: v for k, v in raw.items() if not k.startswith('_')}


# ──────────────────────────────────────────────
# 정규화 함수 레지스트리
# ──────────────────────────────────────────────

def _normalize_age_code(value) -> str:
    """보기개시나이 코드 정규화: '00' -> '0', 공백 → '0'."""
    s = normalize_text(str(value))
    if s == '00':
        return '0'
    return s or '0'


def _normalize_gender(value) -> str:
    """성별 정규화: '1' -> '남자', '2' -> '여자', else ''."""
    s = normalize_text(str(value))
    if s == '1' or s == '남자':
        return '남자'
    if s == '2' or s == '여자':
        return '여자'
    return ''


def _normalize_gender_code(value) -> str:
    s = normalize_text(str(value))
    if s == '1':
        return '1'
    if s == '2':
        return '2'
    return ''


def _norm_age_val(value) -> str:
    """나이/기간 값 정규화: 빈 값 → '', 숫자 → 문자열."""
    s = normalize_text(str(value))
    if not s or s == 'None' or s == 'nan':
        return ''
    return s


NORMALIZERS = {
    'text': lambda v: normalize_text(str(v)),
    'code': lambda v: normalize_code(v),
    'age_code': _normalize_age_code,
    'gender': _normalize_gender,
    'gender_code': _normalize_gender_code,
    'age_val': _norm_age_val,
}


# ──────────────────────────────────────────────
# 범용 튜플 추출
# ──────────────────────────────────────────────

def _should_skip(values: list, fields: list, skip_rule: dict) -> bool:
    """skip_rule에 따라 이 튜플을 건너뛸지 판단."""
    if not skip_rule:
        return False
    mode = skip_rule.get('mode', 'all_empty')
    indices = skip_rule.get('fields', [])
    if not indices:
        return False

    if mode == 'any_empty':
        # 지정 필드 중 하나라도 비어있으면 skip
        return any(not values[i] for i in indices)
    elif mode == 'all_empty':
        # 지정 필드 모두 비어있으면 skip
        return all(not values[i] for i in indices)
    elif mode == 'all_default':
        # 지정 필드 모두 default 값이면 skip
        return all(
            values[i] == fields[i].get('default', '')
            for i in indices
        )
    return False


def _extract_generic_set(data_list: list, ds_config: dict) -> set:
    """추출 JSON 데이터에서 비교용 튜플 세트 생성."""
    fields = ds_config['tuple_fields']
    skip_rule = ds_config.get('skip_rule', {})
    result = set()
    for item in data_list:
        values = []
        for f in fields:
            raw = item.get(f['json'], '')
            norm_fn = NORMALIZERS.get(f.get('norm', 'text'), NORMALIZERS['text'])
            values.append(norm_fn(raw))
        if _should_skip(values, fields, skip_rule):
            continue
        # default 적용
        for i, f in enumerate(fields):
            if not values[i] and 'default' in f:
                values[i] = f['default']
        result.add(tuple(values))
    return result


def _extract_generic_answer_set(answer_rows: list, ds_config: dict) -> set:
    """정답 데이터에서 비교용 튜플 세트 생성."""
    fields = ds_config['tuple_fields']
    skip_rule = ds_config.get('skip_rule', {})
    result = set()
    for r in answer_rows:
        values = []
        for f in fields:
            raw = r.get(f['csv'], '')
            norm_fn = NORMALIZERS.get(f.get('norm', 'text'), NORMALIZERS['text'])
            values.append(norm_fn(raw))
        if _should_skip(values, fields, skip_rule):
            continue
        for i, f in enumerate(fields):
            if not values[i] and 'default' in f:
                values[i] = f['default']
        result.add(tuple(values))
    return result


# ──────────────────────────────────────────────
# 특수 규칙 핸들러
# ──────────────────────────────────────────────


def _apply_period_strip_fallback(mapped_set: set, answer_set: set) -> set:
    """기간 제거 fallback: 정답이 기간정보 없는 경우 추출 결과에서 기간 제거."""
    if not mapped_set or not answer_set or mapped_set == answer_set:
        return mapped_set
    # 12-field 튜플 전용 (join_age)
    sample = next(iter(answer_set))
    if len(sample) < 10:
        return mapped_set
    # 인덱스 3, 6, 9 = min_isrn, min_paym, min_spin
    answer_has_periods = any(
        any(a[i] for i in (3, 6, 9) if i < len(a))
        for a in answer_set
    )
    if answer_has_periods:
        return mapped_set
    stripped = set()
    for t in mapped_set:
        stripped.add(t[:3] + ('',) * (len(t) - 3))
    if stripped == answer_set:
        return stripped
    return mapped_set


# ──────────────────────────────────────────────
# 범용 비교 함수
# ──────────────────────────────────────────────

def generic_compare(mapped_row: dict, answer_rows: list, ds_config: dict) -> dict:
    """범용 비교: 추출 JSON ↔ 정답 행 세트 비교."""
    data_field = ds_config['data_field']
    label = ds_config.get('output_label', 'values')
    data_list = mapped_row.get(data_field, [])

    if not data_list:
        answer_set = _extract_generic_answer_set(answer_rows, ds_config)
        return {
            'matched': False,
            'reason': f'No {data_field} data in mapped row',
            f'mapped_{label}': [],
            f'answer_{label}': sorted(answer_set),
        }

    mapped_set = _extract_generic_set(data_list, ds_config)
    answer_set = _extract_generic_answer_set(answer_rows, ds_config)

    if mapped_set == answer_set:
        return {
            'matched': True,
            'reason': 'Perfect match',
            f'mapped_{label}': sorted(mapped_set),
            f'answer_{label}': sorted(answer_set),
        }
    else:
        return {
            'matched': False,
            'reason': f'{data_field} mismatch',
            f'mapped_{label}': sorted(mapped_set),
            f'answer_{label}': sorted(answer_set),
            'missing_in_mapped': sorted(answer_set - mapped_set),
            'extra_in_mapped': sorted(mapped_set - answer_set),
        }


# ──────────────────────────────────────────────
# 범용 리포트 함수
# ──────────────────────────────────────────────

def generic_answer_report(
    answer_data: Dict[Tuple, List[dict]],
    mapped_rows: List[dict],
    csv_rows: List[dict],
    ds_config: dict,
) -> dict:
    """CSV row (261건) 기준 범용 리포트."""
    data_field = ds_config['data_field']
    label = ds_config.get('output_label', 'values')
    special_rules = ds_config.get('special_rules', [])
    not_supported = set(ds_config.get('not_supported_dtcds', []))

    # 매핑 결과 인덱싱: (dtcd, itcd, prod_dtcd, prod_itcd) → set of tuples
    mapped_index_4: Dict[Tuple[str, str, str, str], set] = {}
    mapped_index_2: Dict[Tuple[str, str], set] = {}
    for row in mapped_rows:
        dtcd = normalize_code(row.get('isrn_kind_dtcd', ''))
        itcd = normalize_code(row.get('isrn_kind_itcd', ''))
        if not (dtcd and itcd):
            continue
        prod_dtcd = normalize_code(row.get('prod_dtcd', ''))
        prod_itcd = normalize_code(row.get('prod_itcd', ''))
        tuples = _extract_generic_set(row.get(data_field, []), ds_config)
        if prod_dtcd and prod_itcd:
            key4 = (dtcd, itcd, prod_dtcd, prod_itcd)
            if key4 not in mapped_index_4:
                mapped_index_4[key4] = set()
            mapped_index_4[key4].update(tuples)
        key2 = (dtcd, itcd)
        if key2 not in mapped_index_2:
            mapped_index_2[key2] = set()
        mapped_index_2[key2].update(tuples)

    # 정답 인덱싱: (dtcd, itcd, prod_dtcd, prod_itcd) → set of tuples
    answer_index: Dict[Tuple[str, str, str, str], set] = {}
    for key_tuple, rows in answer_data.items():
        for r in rows:
            prod_dtcd = normalize_code(r.get('PROD_DTCD', ''))
            prod_itcd = normalize_code(r.get('PROD_ITCD', ''))
            key4 = (key_tuple[0], key_tuple[1], prod_dtcd, prod_itcd)
            if key4 not in answer_index:
                answer_index[key4] = set()
            answer_index[key4].update(_extract_generic_answer_set([r], ds_config))

    rows_out = []
    stats = {
        'total': 0, 'matched': 0, 'unmatched': 0,
        'mismatched': 0, 'no_answer': 0, 'not_supported': 0,
    }

    for csv_row in csv_rows:
        dtcd = csv_row['isrn_kind_dtcd']
        itcd = csv_row['isrn_kind_itcd']
        prod_dtcd = csv_row['prod_dtcd']
        prod_itcd = csv_row['prod_itcd']
        sale_nm = csv_row['isrn_kind_sale_nm']

        # not_supported 처리
        if not_supported and dtcd in not_supported:
            status = 'not_supported'
            stats['not_supported'] += 1
            stats['total'] += 1
            rows_out.append({
                'isrn_kind_dtcd': dtcd,
                'isrn_kind_itcd': itcd,
                'prod_dtcd': prod_dtcd,
                'prod_itcd': prod_itcd,
                'isrn_kind_sale_nm': sale_nm,
                'status': status,
                'note': '사업방법서 미기재 - 추출 불가 상품',
            })
            continue

        key4 = (dtcd, itcd, prod_dtcd, prod_itcd)
        answer_vals = answer_index.get(key4)
        mapped_vals = mapped_index_4.get(key4) or mapped_index_2.get((dtcd, itcd))

        # 특수 규칙 적용
        if 'period_strip_fallback' in special_rules and mapped_vals and answer_vals:
            mapped_vals = _apply_period_strip_fallback(mapped_vals, answer_vals)

        if answer_vals is None:
            status = 'no_answer'
            stats['no_answer'] += 1
        elif mapped_vals is None:
            status = 'unmatched'
            stats['unmatched'] += 1
        elif mapped_vals == answer_vals:
            status = 'matched'
            stats['matched'] += 1
        else:
            status = 'mismatched'
            stats['mismatched'] += 1
        stats['total'] += 1

        entry = {
            'isrn_kind_dtcd': dtcd,
            'isrn_kind_itcd': itcd,
            'prod_dtcd': prod_dtcd,
            'prod_itcd': prod_itcd,
            'isrn_kind_sale_nm': sale_nm,
            'status': status,
        }
        if answer_vals is not None:
            entry[f'answer_{label}'] = sorted([list(a) for a in answer_vals])
        if mapped_vals is not None:
            entry[f'mapped_{label}'] = sorted([list(a) for a in mapped_vals])
        if status == 'mismatched' and mapped_vals and answer_vals:
            missing = answer_vals - mapped_vals
            extra = mapped_vals - answer_vals
            if missing:
                entry['missing_in_mapped'] = sorted([list(a) for a in missing])
            if extra:
                entry['extra_in_mapped'] = sorted([list(a) for a in extra])

        rows_out.append(entry)

    order = {'unmatched': 0, 'mismatched': 1, 'no_answer': 2, 'matched': 3, 'not_supported': 4}
    rows_out.sort(key=lambda r: (order.get(r['status'], 9), r['isrn_kind_dtcd'], r['isrn_kind_itcd']))

    has_answer = stats['matched'] + stats['unmatched'] + stats['mismatched']
    summary = {
        'total_csv_rows': stats['total'],
        'matched': stats['matched'],
        'unmatched': stats['unmatched'],
        'mismatched': stats['mismatched'],
        'no_answer': stats['no_answer'],
        'match_rate_with_answer': f"{stats['matched']/has_answer*100:.1f}%" if has_answer else '0%',
        'match_rate_total': f"{stats['matched']/stats['total']*100:.1f}%" if stats['total'] else '0%',
    }
    if stats['not_supported']:
        summary['not_supported'] = stats['not_supported']

    return {
        'summary': summary,
        'rows': rows_out,
    }


# ──────────────────────────────────────────────
# data-set 설정
# ──────────────────────────────────────────────

@dataclass
class DataSetConfig:
    name: str
    mapped_dir: Path
    answer_excel: Path
    mapping_csv: Path
    report_dir: Path
    file_prefix: str
    answer_key_cols: List[str]
    answer_value_cols: List[str]
    compare_fn: Callable
    answer_report_fn: Callable
    answer_csv: Path = None


def get_dataset_config(name: str) -> DataSetConfig:
    configs = _load_dataset_configs()
    if name not in configs:
        raise ValueError(f"Unknown data-set: {name}. Available: {list(configs.keys())}")

    cfg = configs[name]

    def _compare(mapped_row, answer_rows):
        return generic_compare(mapped_row, answer_rows, cfg)

    def _report(answer_data, mapped_rows, csv_rows):
        return generic_answer_report(answer_data, mapped_rows, csv_rows, cfg)

    answer_csv_path = PROJECT_ROOT / cfg['answer_csv'] if cfg.get('answer_csv') else None

    return DataSetConfig(
        name=name,
        mapped_dir=PROJECT_ROOT / cfg['mapped_dir'],
        answer_excel=PROJECT_ROOT / cfg['answer_excel'],
        answer_csv=answer_csv_path,
        mapping_csv=DEFAULT_MAPPING_CSV,
        report_dir=PROJECT_ROOT / cfg['report_dir'],
        file_prefix=cfg['file_prefix'],
        answer_key_cols=cfg['answer_key_cols'],
        answer_value_cols=cfg['answer_value_cols'],
        compare_fn=_compare,
        answer_report_fn=_report,
    )


# ──────────────────────────────────────────────
# 공통 리포트 생성
# ──────────────────────────────────────────────

def generate_extraction_report(comparison_results: List[dict]) -> dict:
    """추출 데이터 기준 리포트."""
    total = len(comparison_results)
    matched = sum(1 for r in comparison_results if r['comparison']['matched'])
    unmatched = total - matched

    failure_reasons: Dict[str, int] = {}
    for result in comparison_results:
        if not result['comparison']['matched']:
            reason = result['comparison']['reason']
            failure_reasons[reason] = failure_reasons.get(reason, 0) + 1

    file_stats: Dict[str, Dict[str, int]] = {}
    for result in comparison_results:
        fn = result.get('source_file', 'unknown')
        if fn not in file_stats:
            file_stats[fn] = {'total': 0, 'matched': 0, 'unmatched': 0}
        file_stats[fn]['total'] += 1
        if result['comparison']['matched']:
            file_stats[fn]['matched'] += 1
        else:
            file_stats[fn]['unmatched'] += 1

    return {
        'summary': {
            'total_rows': total,
            'matched_rows': matched,
            'unmatched_rows': unmatched,
            'match_rate': f"{matched/total*100:.1f}%" if total else '0%',
        },
        'failure_reasons': failure_reasons,
        'file_statistics': [
            {
                'file': fn,
                'total': s['total'],
                'matched': s['matched'],
                'unmatched': s['unmatched'],
                'match_rate': f"{s['matched']/s['total']*100:.1f}%" if s['total'] else '0%',
            }
            for fn, s in sorted(file_stats.items())
        ],
        'sample_failures': [
            {
                'source_file': r['source_file'],
                'product': r.get('상품명', ''),
                'isrn_kind_dtcd': r.get('isrn_kind_dtcd', ''),
                'isrn_kind_itcd': r.get('isrn_kind_itcd', ''),
                'reason': r['comparison']['reason'],
                'mapped': r['comparison'].get(f'mapped_{list(r["comparison"].keys())[0]}', []),
                'answer': r['comparison'].get(f'answer_{list(r["comparison"].keys())[0]}', []),
                'missing': r['comparison'].get('missing_in_mapped', []),
                'extra': r['comparison'].get('extra_in_mapped', []),
            }
            for r in comparison_results[:20]
            if not r['comparison']['matched']
        ],
    }


# ──────────────────────────────────────────────
# 메인 비교 루프
# ──────────────────────────────────────────────

def run_comparison(config: DataSetConfig, verbose: bool = False):
    print(f"[{config.name}] Loading CSV mapping...")
    csv_rows = load_csv_rows(config.mapping_csv)
    print(f"  {len(csv_rows)} CSV rows")

    if config.answer_csv and config.answer_csv.exists():
        print(f"[{config.name}] Loading answer CSV: {config.answer_csv.name}")
        answer_data = load_answer_csv(
            config.answer_csv, config.answer_key_cols, config.answer_value_cols,
        )
    else:
        print(f"[{config.name}] Loading answer Excel...")
        answer_data = load_answer_excel(
            config.answer_excel, config.answer_key_cols, config.answer_value_cols,
        )
    print(f"  {sum(len(v) for v in answer_data.values())} answer rows, "
          f"{len(answer_data)} unique products")

    print(f"\n[{config.name}] Loading mapped JSON files...")
    mapped_rows = load_mapped_json_files(config.mapped_dir, config.file_prefix)
    print(f"  {len(mapped_rows)} mapped rows")

    print(f"\n[{config.name}] Comparing...")
    comparison_results = []

    for mapped_row in mapped_rows:
        dtcd = normalize_code(mapped_row.get('isrn_kind_dtcd', ''))
        itcd = normalize_code(mapped_row.get('isrn_kind_itcd', ''))
        sale_nm = normalize_text(str(mapped_row.get('isrn_kind_sale_nm', '')))
        prod_dtcd = normalize_code(mapped_row.get('prod_dtcd', ''))
        prod_itcd = normalize_code(mapped_row.get('prod_itcd', ''))

        if not (dtcd and itcd and sale_nm):
            comparison_results.append({
                'source_file': mapped_row.get('source_file', ''),
                'isrn_kind_dtcd': dtcd,
                'isrn_kind_itcd': itcd,
                'isrn_kind_sale_nm': sale_nm,
                'prod_dtcd': prod_dtcd,
                'prod_itcd': prod_itcd,
                '상품명': mapped_row.get('상품명', ''),
                'comparison': {
                    'matched': False,
                    'reason': 'No mapping found (empty codes)',
                    'mapped_cycles': [],
                    'answer_cycles': [],
                },
            })
            continue

        key = tuple(
            normalize_code(v) if c in ('ISRN_KIND_DTCD', 'ISRN_KIND_ITCD')
            else normalize_text(str(v))
            for c, v in zip(config.answer_key_cols, [dtcd, itcd, sale_nm])
        )
        answer_rows = answer_data.get(key, [])

        # prod 코드로 정답 행 필터링 (적립형/거치형 등 변종 분리)
        if prod_dtcd and prod_itcd and answer_rows:
            filtered = [
                r for r in answer_rows
                if normalize_code(r.get('PROD_DTCD', '')) == prod_dtcd
                and normalize_code(r.get('PROD_ITCD', '')) == prod_itcd
            ]
            if filtered:
                answer_rows = filtered

        if not answer_rows:
            comparison_results.append({
                'source_file': mapped_row.get('source_file', ''),
                'isrn_kind_dtcd': dtcd,
                'isrn_kind_itcd': itcd,
                'isrn_kind_sale_nm': sale_nm,
                'prod_dtcd': prod_dtcd,
                'prod_itcd': prod_itcd,
                '상품명': mapped_row.get('상품명', ''),
                'comparison': {
                    'matched': False,
                    'reason': 'Product not found in answer',
                    'mapped_cycles': [],
                    'answer_cycles': [],
                },
            })
            continue

        comparison = config.compare_fn(mapped_row, answer_rows)
        comparison_results.append({
            'source_file': mapped_row.get('source_file', ''),
            'isrn_kind_dtcd': dtcd,
            'isrn_kind_itcd': itcd,
            'isrn_kind_sale_nm': sale_nm,
            'prod_dtcd': prod_dtcd,
            'prod_itcd': prod_itcd,
            '상품명': mapped_row.get('상품명', ''),
            'comparison': comparison,
        })

        if verbose and not comparison['matched']:
            print(f"\n  Mismatch: {mapped_row.get('source_file', '')}")
            print(f"    Product: {mapped_row.get('상품명', '')}")

    # ── 리포트 저장 ──
    config.report_dir.mkdir(parents=True, exist_ok=True)

    # 1) 추출 기준 리포트
    report = generate_extraction_report(comparison_results)
    report_path = config.report_dir / 'comparison_report.json'
    with report_path.open('w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    # 2) 전체 비교 상세
    detailed_path = config.report_dir / 'comparison_detailed.json'
    with detailed_path.open('w', encoding='utf-8') as f:
        json.dump(comparison_results, f, ensure_ascii=False, indent=2)

    # 3) CSV row 기준 리포트
    answer_report = config.answer_report_fn(answer_data, mapped_rows, csv_rows)
    answer_report_path = config.report_dir / 'answer_based_report.json'
    with answer_report_path.open('w', encoding='utf-8') as f:
        json.dump(answer_report, f, ensure_ascii=False, indent=2)

    # ── 콘솔 출력 ──
    print(f"\n{'='*50}")
    print(f"EXTRACTION-BASED REPORT ({config.name})")
    print(f"{'='*50}")
    s = report['summary']
    print(f"Total rows: {s['total_rows']}")
    print(f"Matched:    {s['matched_rows']}")
    print(f"Unmatched:  {s['unmatched_rows']}")
    print(f"Match rate: {s['match_rate']}")
    if report['failure_reasons']:
        print("\nFailure reasons:")
        for reason, count in sorted(report['failure_reasons'].items(), key=lambda x: -x[1]):
            print(f"  {reason}: {count}")
    print(f"\nReport: {report_path}")
    print(f"Detail: {detailed_path}")

    s2 = answer_report['summary']
    print(f"\n{'='*50}")
    print(f"CSV-ROW-BASED REPORT ({config.name})")
    print(f"{'='*50}")
    print(f"Total CSV rows: {s2['total_csv_rows']}")
    print(f"  Matched:    {s2['matched']}")
    print(f"  Unmatched:  {s2['unmatched']} (매핑 데이터 없음)")
    print(f"  Mismatched: {s2['mismatched']} (값 불일치)")
    print(f"  No answer:  {s2['no_answer']} (정답 없음)")
    print(f"  Match rate (정답 있는 건): {s2['match_rate_with_answer']}")
    print(f"  Match rate (전체):        {s2['match_rate_total']}")
    print(f"Report: {answer_report_path}")


# ──────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────

def run_single_comparison(config: DataSetConfig, json_path: Path,
                          output_path: Optional[Path] = None,
                          verbose: bool = False):
    """단일 매핑 JSON 파일에 대한 비교."""
    # 정답 로드
    if config.answer_csv and config.answer_csv.exists():
        answer_data = load_answer_csv(
            config.answer_csv, config.answer_key_cols, config.answer_value_cols,
        )
    else:
        answer_data = load_answer_excel(
            config.answer_excel, config.answer_key_cols, config.answer_value_cols,
        )

    # 단일 파일 로드
    with json_path.open('r', encoding='utf-8') as f:
        data = json.load(f)
    mapped_rows = []
    if isinstance(data, list):
        for row in data:
            if isinstance(row, dict):
                row['source_file'] = json_path.name
                mapped_rows.append(row)

    print(f"[{config.name}] Single file: {json_path.name}")
    print(f"  {len(mapped_rows)} mapped rows, "
          f"{sum(len(v) for v in answer_data.values())} answer rows")

    # 비교 루프 (run_comparison과 동일)
    comparison_results = []
    stats = {'matched': 0, 'unmatched': 0, 'no_answer': 0, 'no_mapping': 0}

    for mapped_row in mapped_rows:
        dtcd = normalize_code(mapped_row.get('isrn_kind_dtcd', ''))
        itcd = normalize_code(mapped_row.get('isrn_kind_itcd', ''))
        sale_nm = normalize_text(str(mapped_row.get('isrn_kind_sale_nm', '')))
        prod_dtcd = normalize_code(mapped_row.get('prod_dtcd', ''))
        prod_itcd = normalize_code(mapped_row.get('prod_itcd', ''))

        result_base = {
            'source_file': mapped_row.get('source_file', ''),
            'isrn_kind_dtcd': dtcd,
            'isrn_kind_itcd': itcd,
            'isrn_kind_sale_nm': sale_nm,
            'prod_dtcd': prod_dtcd,
            'prod_itcd': prod_itcd,
            '상품명': mapped_row.get('상품명', ''),
        }

        if not (dtcd and itcd and sale_nm):
            result_base['comparison'] = {
                'matched': False, 'reason': 'No mapping found (empty codes)',
            }
            stats['no_mapping'] += 1
            comparison_results.append(result_base)
            continue

        key = tuple(
            normalize_code(v) if c in ('ISRN_KIND_DTCD', 'ISRN_KIND_ITCD')
            else normalize_text(str(v))
            for c, v in zip(config.answer_key_cols, [dtcd, itcd, sale_nm])
        )
        answer_rows = answer_data.get(key, [])

        if prod_dtcd and prod_itcd and answer_rows:
            filtered = [
                r for r in answer_rows
                if normalize_code(r.get('PROD_DTCD', '')) == prod_dtcd
                and normalize_code(r.get('PROD_ITCD', '')) == prod_itcd
            ]
            if filtered:
                answer_rows = filtered

        if not answer_rows:
            result_base['comparison'] = {
                'matched': False, 'reason': 'Product not found in answer',
            }
            stats['no_answer'] += 1
            comparison_results.append(result_base)
            continue

        comparison = config.compare_fn(mapped_row, answer_rows)
        result_base['comparison'] = comparison
        if comparison['matched']:
            stats['matched'] += 1
        else:
            stats['unmatched'] += 1
        comparison_results.append(result_base)

        if verbose and not comparison['matched']:
            print(f"  Mismatch: {mapped_row.get('상품명', '')}")

    # 출력
    total = len(comparison_results)
    if output_path is None:
        output_path = json_path.with_suffix('.compare.json')
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open('w', encoding='utf-8') as f:
        json.dump(comparison_results, f, ensure_ascii=False, indent=2)

    print(f"\n  Total: {total}")
    print(f"  Matched: {stats['matched']}")
    print(f"  Mismatched: {stats['unmatched']}")
    print(f"  No answer: {stats['no_answer']}")
    print(f"  No mapping: {stats['no_mapping']}")
    if total:
        rate = stats['matched'] / total * 100
        print(f"  Match rate: {rate:.1f}%")
    print(f"  Report: {output_path}")


def parse_args():
    configs = _load_dataset_configs()
    parser = argparse.ArgumentParser(description='통합 상품 데이터 비교')
    parser.add_argument('--data-set', required=True,
                        choices=list(configs.keys()),
                        help='비교할 데이터셋')
    parser.add_argument('--json', type=str, default=None,
                        help='단일 매핑 JSON 파일 경로')
    parser.add_argument('--output', type=str, default=None,
                        help='단일 파일 비교 결과 출력 경로')
    parser.add_argument('--answer-excel', type=Path, default=None,
                        help='정답 Excel 경로 (기본: data-set별 자동)')
    parser.add_argument('--mapped-dir', type=Path, default=None,
                        help='매핑 데이터 디렉토리 (기본: data-set별 자동)')
    parser.add_argument('--report-dir', type=Path, default=None,
                        help='리포트 출력 디렉토리 (기본: 정답비교/{data-set})')
    parser.add_argument('--verbose', action='store_true')
    return parser.parse_args()


def main():
    args = parse_args()

    config = get_dataset_config(args.data_set)

    # CLI 오버라이드
    if args.answer_excel:
        config.answer_excel = args.answer_excel

    # --- 단일 파일 모드 ---
    if args.json:
        json_path = Path(args.json)
        output_path = Path(args.output) if args.output else None
        run_single_comparison(config, json_path, output_path, verbose=args.verbose)
        return

    # --- 디렉토리 모드 ---
    if args.mapped_dir:
        config.mapped_dir = args.mapped_dir
    if args.report_dir:
        config.report_dir = args.report_dir

    run_comparison(config, verbose=args.verbose)


if __name__ == '__main__':
    main()
